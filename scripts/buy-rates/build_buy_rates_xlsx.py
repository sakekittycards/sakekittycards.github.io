"""
Branded, shareable Buy-Rates Calculator workbook (.xlsx) for Sake Kitty Cards.

Self-calculating: pick a Category (dropdown), type Market Value + Qty, and the
Cash / Credit offers + per-category subtotals + grand total auto-fill.
Opens in Excel and Google Sheets. Rates mirror trade-in.html (cash +5pp per
2026-05-26 request). Re-run if rates change.
"""
import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

REPO = r"C:\Users\lunar\OneDrive\Desktop\sake-kitty-cards-site"
OUT = os.environ.get("OUT", r"C:\Users\lunar\OneDrive\Desktop\Sake_Kitty_Buy_Rates.xlsx")
CONTACT = "nick@sakekittycards.com"

# ── rates (from trade-in.html; cash +5pp 2026-05-26) ─────────────────────────
TIERS = [(0, 0.70, 0.80), (100, 0.80, 0.90), (500, 0.85, 0.95), (1000, 0.90, 1.00)]
PCT_CATS = ["Singles", "Sealed", "Graded"]
BULK = [
    ("Unsorted Bulk (by weight)", 1.50, 2.50), ("Common / Uncommon", 0.01, 0.01),
    ("Rare (Non-Holo)", 0.02, 0.03), ("Reverse Holo", 0.03, 0.05),
    ("Holo Rare", 0.05, 0.10), ("Holo Promo", 0.10, 0.15), ("EX or V", 0.15, 0.25),
    ("Radiant", 0.20, 0.30), ("GX", 0.35, 0.50), ("VMAX / VSTAR", 0.45, 0.65),
    ("Full Art", 0.50, 0.75), ("Amazing Rare", 0.70, 1.05), ("BREAK", 0.75, 1.10),
]
DATA_FIRST, DATA_LAST = 5, 514                      # 510 entry rows

# ── brand palette ────────────────────────────────────────────────────────────
ORANGE, PINK, PURPLE, CYAN, DARK = "FF6A00", "FF0080", "7B2FFF", "00D4FF", "0B0B14"
WHITE, INK, LIGHT, LINE = "FFFFFF", "1A1A22", "FFF4EC", "E4D9CC"
CUR, PCT = '"$"#,##0.00', '0%'

def hexlum(rgb):  # 0..255 luminance
    r, g, b = rgb[:3]
    return 0.299 * r + 0.587 * g + 0.114 * b

wb = Workbook()
ws = wb.active
ws.title = "Buy Rates"
ws.sheet_view.showGridLines = False

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

def S(cell, val=None, *, font=None, fill=None, align=None, fmt=None, border=None):
    c = ws[cell]
    if val is not None: c.value = val
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = border
    return c

# ── header band (rows 1-3), blended to the logo's background colour ──────────
logo_path = os.path.join(REPO, "logo.png")
band = DARK
text_on_band = WHITE
try:
    im = Image.open(logo_path).convert("RGB")
    corner = im.getpixel((3, 3))
    band = "%02X%02X%02X" % corner
    text_on_band = INK if hexlum(corner) > 150 else WHITE
except Exception:
    pass

# fill the whole band first (every underlying cell, so merges render solid)
for r in (1, 2, 3):
    for col in range(1, 12):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=band)
    ws.row_dimensions[r].height = 26
ws.merge_cells("A1:K2")
S("A1", "SAKE KITTY CARDS", font=Font(name="Arial Black", size=22, bold=True, color=text_on_band),
  align=Alignment(horizontal="center", vertical="center"))
ws.merge_cells("A3:K3")
S("A3", f"Buy Rates Calculator   ·   {CONTACT}",
  font=Font(size=10, italic=True, color=text_on_band),
  align=Alignment(horizontal="center", vertical="center"))

# logo image, top-left over the band
try:
    tmp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_logo_64.png")
    Image.open(logo_path).convert("RGBA").resize((150, 150)).save(tmp)
    xli = XLImage(tmp); xli.width = 66; xli.height = 66
    ws.add_image(xli, "A1")
except Exception as e:
    print("logo embed skipped:", e)

# ── data-entry headers (row 4) ───────────────────────────────────────────────
heads = ["Category", "Item / Description", "Market Value $", "Qty / Lbs", "Cash Offer", "Credit Offer"]
for i, h in enumerate(heads):
    S(f"{get_column_letter(i+1)}4", h, font=Font(bold=True, color=WHITE, size=11),
      fill=ORANGE, align=center, border=box)
ws.row_dimensions[4].height = 22

# ── example rows ─────────────────────────────────────────────────────────────
examples = [
    ("Singles", "EXAMPLE — Charizard (delete row)", 120, 1),
    ("Sealed", "EXAMPLE — Evolving Skies Booster Box", 450, 1),
    ("Holo Rare", "EXAMPLE — bulk holo rares", None, 50),
    ("Unsorted Bulk (by weight)", "EXAMPLE — unsorted pile (lbs in Qty)", None, 10),
]
for i, (cat, item, val, qty) in enumerate(examples):
    r = DATA_FIRST + i
    S(f"A{r}", cat); S(f"B{r}", item)
    if val is not None: S(f"C{r}", val, fmt=CUR)
    S(f"D{r}", qty)

# ── per-row formulas + cell styling for all entry rows ───────────────────────
def cash(r):
    return (f'=IF($A{r}="","",IFERROR(IF(VLOOKUP($A{r},$H$12:$K$27,2,FALSE)="PCT",'
            f'IF(AND(LOWER($A{r})="graded",$C{r}<100),"NOT ACCEPTED (<$100)",'
            f'ROUND($C{r}*VLOOKUP($C{r},$H$6:$J$9,2,TRUE)*IF($D{r}="",1,$D{r}),2)),'
            f'ROUND(VLOOKUP($A{r},$H$12:$K$27,3,FALSE)*IF($D{r}="",1,$D{r}),2)),"check category"))')
def credit(r):
    return (f'=IF($A{r}="","",IFERROR(IF(VLOOKUP($A{r},$H$12:$K$27,2,FALSE)="PCT",'
            f'IF(AND(LOWER($A{r})="graded",$C{r}<100),"NOT ACCEPTED (<$100)",'
            f'ROUND($C{r}*VLOOKUP($C{r},$H$6:$J$9,3,TRUE)*IF($D{r}="",1,$D{r}),2)),'
            f'ROUND(VLOOKUP($A{r},$H$12:$K$27,4,FALSE)*IF($D{r}="",1,$D{r}),2)),"check category"))')
for r in range(DATA_FIRST, DATA_LAST + 1):
    S(f"C{r}", fmt=CUR)
    S(f"E{r}", cash(r), fmt=CUR, font=Font(bold=True, color=ORANGE))
    S(f"F{r}", credit(r), fmt=CUR, font=Font(bold=True, color=PURPLE))

# Category dropdown on column A (foolproof for sharing)
dv = DataValidation(type="list", formula1="=$H$12:$H$27", allow_blank=True, showDropDown=False)
dv.error = "Pick a category from the list."
dv.prompt = "Choose: Singles, Sealed, Graded, or a bulk type"
ws.add_data_validation(dv)
dv.add(f"A{DATA_FIRST}:A{DATA_LAST}")

# ── TIER RATES table (H5 header, data H6:J9) ─────────────────────────────────
S("H4", "TIER RATES — Singles / Sealed / Graded", font=Font(bold=True, color=WHITE), fill=CYAN, align=left)
ws.merge_cells("H4:K4")
for c, h in zip("HIJ", ["Min $", "Cash %", "Credit %"]):
    S(f"{c}5", h, font=Font(bold=True), fill=LIGHT, align=center, border=box)
for i, (mn, ca, cr) in enumerate(TIERS):
    r = 6 + i
    S(f"H{r}", mn, fmt=CUR, align=center, border=box)
    S(f"I{r}", ca, fmt=PCT, align=center, border=box)
    S(f"J{r}", cr, fmt=PCT, align=center, border=box)

# ── CATEGORY table (H11 header, data H12:K27) ────────────────────────────────
S("H10", "CATEGORIES & RATES (edit any number)", font=Font(bold=True, color=WHITE), fill=PURPLE, align=left)
ws.merge_cells("H10:K10")
for c, h in zip("HIJK", ["Category", "Type", "Flat Cash $", "Flat Credit $"]):
    S(f"{c}11", h, font=Font(bold=True), fill=LIGHT, align=center, border=box)
row = 12
for name in PCT_CATS:
    S(f"H{row}", name, align=left, border=box); S(f"I{row}", "PCT", align=center, border=box)
    S(f"J{row}", "(tiered)", align=center, border=box); S(f"K{row}", "(tiered)", align=center, border=box)
    row += 1
for name, ca, cr in BULK:
    S(f"H{row}", name, align=left, border=box); S(f"I{row}", "FLAT", align=center, border=box)
    S(f"J{row}", ca, fmt=CUR, align=center, border=box); S(f"K{row}", cr, fmt=CUR, align=center, border=box)
    row += 1

# ── SUMMARY (H30 title, data H32:K47, grand total H48) ───────────────────────
S("H30", "SUMMARY — per category", font=Font(bold=True, color=WHITE), fill=ORANGE, align=left)
ws.merge_cells("H30:K30")
for c, h in zip("HIJK", ["Category", "Cash Total", "Credit Total", "Lines"]):
    S(f"{c}31", h, font=Font(bold=True), fill=LIGHT, align=center, border=box)
all_cats = PCT_CATS + [b[0] for b in BULK]
sr = 32
for name in all_cats:
    S(f"H{sr}", name, align=left, border=box)
    S(f"I{sr}", f'=ROUND(SUMIF($A$5:$A$514,$H{sr},$E$5:$E$514),2)', fmt=CUR, align=center, border=box)
    S(f"J{sr}", f'=ROUND(SUMIF($A$5:$A$514,$H{sr},$F$5:$F$514),2)', fmt=CUR, align=center, border=box)
    S(f"K{sr}", f'=COUNTIF($A$5:$A$514,$H{sr})', align=center, border=box)
    sr += 1
gt = sr
S(f"H{gt}", "GRAND TOTAL", font=Font(bold=True, color=WHITE), fill=DARK, align=left, border=box)
S(f"I{gt}", '=ROUND(SUM($E$5:$E$514),2)', font=Font(bold=True, color=WHITE), fill=DARK, fmt=CUR, align=center, border=box)
S(f"J{gt}", '=ROUND(SUM($F$5:$F$514),2)', font=Font(bold=True, color=WHITE), fill=DARK, fmt=CUR, align=center, border=box)
S(f"K{gt}", '=COUNTA($A$5:$A$514)', font=Font(bold=True, color=WHITE), fill=DARK, align=center, border=box)

# ── how-to notes ─────────────────────────────────────────────────────────────
notes = [
    "HOW TO USE",
    "1. Pick a Category from the dropdown in column A.",
    "2. Singles / Sealed / Graded: enter Market Value $ (col C). Qty in D (blank = 1).",
    "3. Bulk categories: leave Value blank, put # of cards in Qty (col D).",
    "4. Unsorted Bulk (by weight): put POUNDS in the Qty/Lbs column.",
    "5. Cash & Credit auto-fill. Graded under $100 = not accepted.",
    "6. Edit any number in the tables above and everything recalculates.",
]
nr = gt + 2
S(f"H{nr}", notes[0], font=Font(bold=True, color=PINK))
for i, line in enumerate(notes[1:], 1):
    S(f"H{nr+i}", line, font=Font(size=10, color=INK))

# ── widths, freeze, formats ──────────────────────────────────────────────────
widths = {"A": 22, "B": 34, "C": 14, "D": 10, "E": 13, "F": 13, "G": 3,
          "H": 30, "I": 12, "J": 12, "K": 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

wb.save(OUT)
print("wrote", OUT)
